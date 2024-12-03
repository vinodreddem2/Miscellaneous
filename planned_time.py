import argparse
import logging
import json
import pandas as pd
from collections import defaultdict
import openpyxl
import requests
logging.basicConfig(level=logging.DEBUG)


jira_data_with_epic_points = {
    'VWICAS23-232204': {  # Epic 1
        'Kiran': {'planned': 0.74375},
        'Srinivas': {'planned': 0.95625},
        'Rishika': {'planned': 0.74375},
        'Jay': {'planned': 0.53125},
        'Giridhar': {'planned': 0.725},
        'Vijaya': {'planned': 0.8125},
        'Abishek': {'planned': 0.71875},
        'Gajanan': {'planned': 0.875},
        'Gopika': {'planned': 0.7375},
        'Elango': {'planned': 0.775},
        'Adil': {'planned': 0.75},
    },
    'VWICAS23-232205': {  # Epic 2
        'Kiran': {'planned': 1.25},
        'Srinivas': {'planned': 0.095625},
        'Rishika': {'planned': 0.19125},
        'Jay': {'planned': 0.0425},
        'Giridhar': {'planned': 0.25},
        'Vijaya': {'planned': 0.5375},
        'Abishek': {'planned': 0.4625},
        'Gajanan': {'planned': 0.6375},
        'Gopika': {'planned': 0.625},
        'Elango': {'planned': 0.2625},
        'Adil': {'planned': 0.375},
    },
    'VWICAS23-232206': {  # Epic 3
        'Kiran': {'planned': 0.5},
        'Srinivas': {'planned': 0.625},
        'Rishika': {'planned': 0.25},
        'Jay': {'planned': 0.375},
        'Giridhar': {'planned': 0.3125},
        'Vijaya': {'planned': 0.5},
        'Abishek': {'planned': 0.4375},
        'Gajanan': {'planned': 0.625},
        'Gopika': {'planned': 0.3125},
        'Elango': {'planned': 0.625},
        'Adil': {'planned': 0.375},
    },
    'VWICAS23-232207': {  # Epic 4
        'Kiran': {'planned': 0.95},
        'Srinivas': {'planned': 1.2},
        'Rishika': {'planned': 0.4},
        'Jay': {'planned': 0.325},
        'Giridhar': {'planned': 0.45},
        'Vijaya': {'planned': 0.6},
        'Abishek': {'planned': 0.725},
        'Gajanan': {'planned': 0.875},
        'Gopika': {'planned': 0.75},
        'Elango': {'planned': 0.4},
        'Adil': {'planned': 0.45},
    },
    'VWICAS23-232208': {  # Epic 5
        'Kiran': {'planned': 0.85},
        'Srinivas': {'planned': 0.7},
        'Rishika': {'planned': 1.25},
        'Jay': {'planned': 0.55},
        'Giridhar': {'planned': 0.9},
        'Vijaya': {'planned': 1.0},
        'Abishek': {'planned': 0.65},
        'Gajanan': {'planned': 1.0},
        'Gopika': {'planned': 1.25},
        'Elango': {'planned': 0.875},
        'Adil': {'planned': 0.9},
    }
}


def get_ticket_details(ticket_key):
    # Define the custom field and Jira base URL
    custom_field_clause_name = 'cf[13504]'  # Replace with your custom field
    jira_base_url = 'https://jira-ibs.zone2.agileci.conti.de'
    url = f"{jira_base_url}/rest/api/2/search"

    # Load API_KEY from config file
    with open('config.json') as config_file:
        config_data = json.load(config_file)
        API_KEY = config_data.get('API_KEY', '')  # Make sure to set API_KEY in your config file
    
    headers = {
        "Accept": "application/json",
        'Content-Type': 'application/json',
        "Authorization": f"Bearer {API_KEY}"  # Authorization header for API access
    }

    # JQL query to fetch issues based on specific ticket key (e.g., 'VWICAS23-232204')
    jql_query = f'key = "{ticket_key}"'

    # Make the API request to Jira
    params = {
        'jql': jql_query,
        'fields': 'summary,issuetype,customfield_13504',  # Specify the fields you need
        'maxResults': 1  # We only want to fetch one issue (based on ticket_key)
    }

    # Send the GET request
    response = requests.get(url, headers=headers, params=params)

    if response.status_code == 200:        
        data = response.json()
        issues = data.get('issues', [])
        if issues:
            issue = issues[0]
            summary = issue['fields']['summary']  # Summary of the ticket
            jira_url = f"{jira_base_url}/browse/{ticket_key}"  # Jira link URL
            print(f"Ticket Summary: {summary}")
            print(f"Jira URL: {jira_url}")
            return summary, jira_url
        else:
            print(f"No issue found for ticket key: {ticket_key}")
    else:        
        print(f"Error fetching ticket details: {response.status_code} - {response.text}")
        return None, None
  

def fetch_estimates_and_time_spent(issue_type, responsible_team, sprint_id):    
    custom_field_clause_name = 'cf[13504]'  # Replace with your custom field
    jira_base_url = 'https://jira-ibs.zone2.agileci.conti.de'
    url = f"{jira_base_url}/rest/api/2/search"

    with open('config.json') as config_file:
        config_data = json.load(config_file)
        API_KEY = config_data.get('API_KEY', '')
    
    headers = {
        "Accept": "application/json",
        'Content-Type': 'application/json',
        "Authorization": API_KEY
    }
    
    # JQL query to fetch issues based on issue type, team, and sprint
    jql_query = f'issuetype = "{issue_type}" AND "{custom_field_clause_name}" = "{responsible_team}" AND Sprint = {sprint_id}'
    
    start_at = 0
    max_results = 50
    total_results = 1  # Set initially to enter the loop
    grouped_data = defaultdict(lambda: defaultdict(float))

    # Loop through paginated results
    while start_at < total_results:
        params = {
            "jql": jql_query,
            "fields": "summary,key,assignee,timeoriginalestimate,timespent,customfield_10000, subtasks",
            "startAt": start_at,
            "maxResults": max_results
        }
        
        response = requests.get(url, headers=headers, params=params)
        logging.debug(f"Request URL: {response.url}")

        if response.status_code == 200:
            data = response.json()
            total_results = data['total']  # Total number of issues matching the query
            issues = data.get('issues', [])

            if not issues and start_at == 0:
                logging.error("No issues found matching the JQL query.")
                return []

            for issue in issues:
                customfield = issue['fields']['customfield_10000']  # Epic or custom field ID
                assignee_name = issue['fields']['assignee']['displayName']  # Assignee name
                time_estimate = issue['fields'].get('timeoriginalestimate', 0)  # Time estimate (defaults to 0 if missing)

                if issue.get('fields', {}).get('subtasks'):
                    for subtask in issue['fields']['subtasks']:
                        for subtask in issue['fields']['subtasks']:
                            pass
                if time_estimate is None:
                    time_estimate = 0
                seconds_per_story_point = 8 * 60 * 60
                story_points = time_estimate / seconds_per_story_point
                grouped_data[customfield][assignee_name] += story_points
            start_at += max_results
            print(f"start_at  --- {start_at}")
            print(f"total_results  --- {total_results}")
                
    return grouped_data

# Function to convert the data into the desired format
def convert_to_epic_points(mydata):
    result = {}
    
    for epic, assignees in mydata.items():
        result[epic] = {}
        
        for assignee, hours in assignees.items():
            epic_points = round(hours / 8, 2)
            result[epic][assignee] = {'planned': epic_points}
            
    return result


def update_planned_hours_in_excel(excel_file, jira_data, target_sheet_name):    
    wb = openpyxl.load_workbook(excel_file)
       # Get the specific sheet by name
    if target_sheet_name in wb.sheetnames:
        sheet = wb[target_sheet_name]
    else:
        print(f"Sheet '{target_sheet_name}' not found in the workbook.")
        return

    assignees = []
    header_row = sheet[1]
    
    for col in range(2, len(header_row)):
        assignee_name = header_row[col].value
        if assignee_name and assignee_name != "Key" and assignee_name != "Summary":
            assignees.append(assignee_name)
    
    print(f"Assignees found in the sheet: {assignees}")

    # Iterate over each row in the sheet
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        epic_link = row[0].value  # Epic link in the first column
        epic_name = row[1].value  # Epic name in the second column
        if not epic_link or not epic_name:
            # print(f"Skipping row: Epic link or Epic name is missing. epic_link: {epic_link}, epic_name: {epic_name}")
            continue
        print('#'*150)
        print("epic_name is ", epic_name)
        print('epic_link is ', epic_link)
        
        # Fetch the JIRA data for the epic link (if available)
        epic_data = jira_data.get(epic_name, {})
        
        if epic_data:
            # Iterate through each assignee and update their "Planned" column
            for idx, assignee in enumerate(assignees, start=0):  # Start from the second column (first assignee)
                if assignee in epic_data:
                    print("Assignee data is found ", assignee)
                    planned_hours = epic_data[assignee].get('planned', 0)
                    print("planned_hours is ", planned_hours)
                    # Calculate the correct column for the "Planned" based on the index
                    planned_column = 3 + (idx * 3) + 2 
                    print('Cell Value is ', row[0].row, planned_column)
                    sheet.cell(row=row[0].row, column=planned_column, value=planned_hours)
        else:
            # If no data found for the epic, continue without making changes to the "Planned" columns
            print(f"No data found for epic {epic_link} (Epic name: {epic_name}), skipping update.")
    
    # Save the updated workbook
    wb.save(excel_file)


def main():
    parser = argparse.ArgumentParser(description="Fetch Jira estimates and time spent for each team member.")
    parser.add_argument('--issue_type', help="Issue type (e.g., Sub-task)", required=True)
    parser.add_argument('--responsible_team', help="Responsible SAFe Team", required=True)
    parser.add_argument('--sprint_id', help="Sprint ID", required=True)
    
    args = parser.parse_args()

    # Fetch estimates and time spent from Jira
    processed_data = fetch_estimates_and_time_spent(args.issue_type, args.responsible_team, args.sprint_id)    
    processed_data = convert_to_epic_points(processed_data)
 
    update_planned_hours_in_excel('PIPE_Epic_Planning.xlsx', jira_data_with_epic_points)



def RunServer(logger):
    try:
        logger.info("Starting server...")
        asgi_app = get_asgi_application()
        logger.info("ASGI app loaded successfully")

        https_settings = get_https_settings(logger)
        ssl_keyfile = https_settings.get('ssl_keyfile')
        ssl_certfile = https_settings.get('ssl_certfile')

        logger.info(f"ssl_keyfile is: {ssl_keyfile}")
        logger.info(f"ssl_certfile is {ssl_certfile}")
        is_http_server_started = False

        if ssl_keyfile and ssl_certfile:
            if os.path.exists(ssl_keyfile) and os.path.exists(ssl_certfile):
                pass
            else:
                logger.warning("Invalid SSL key and SSL cert file paths provided...")
        else:
            logger.info("No SSL key and/or cert found skipping HTTPS setup.")

        if not is_http_server_started:            
            logger.info("Starting HTTP server")
            pass
    except Exception as _:
        LogException('RunServer', logger)
        sys.exit(1)

if __name__ == "__main__":
    main()


