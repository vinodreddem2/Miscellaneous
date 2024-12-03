import argparse
import logging
import json
import pandas as pd
from collections import defaultdict
import openpyxl
import requests

logging.basicConfig(level=logging.DEBUG)


def get_ticket_details(ticket_key):

    url = f"https://jira-ibs.zone2.agileci.conti.de:443/rest/api/2/issue/{ticket_key.strip()}"

    with open('config.json') as config_file:
        config_data = json.load(config_file)
        API_KEY = config_data.get('API_KEY', '')  # Make sure to set API_KEY in your config file

    payload = {}
    headers = { "Accept": "application/json",
    'Content-Type': 'application/json',
    'Authorization': API_KEY
    }
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    return None


def fetch_estimates_and_time_spent(issue_type, responsible_team, label):
    custom_field_clause_name = 'cf[13504]'  # Replace with your custom field
    epic_field = 'customfield_10000'  # Epic field
    jira_base_url = 'https://jira-ibs.zone2.agileci.conti.de'
    url = f"{jira_base_url}/rest/api/2/search"

    with open('config.json') as config_file:
        config_data = json.load(config_file)
        API_KEY = config_data.get('API_KEY', '')
    
    headers = {
        "Accept": "application/json",
        'Content-Type': 'application/json',
        "Authorization": API_KEY
    }
        
    jql_query = f'issuetype = "{issue_type}" AND "{custom_field_clause_name}" = "{responsible_team}" AND labels = {label}'
    
    start_at = 0
    max_results = 50
    total_results = 1  # Set initially to enter the loop
    team_data =  defaultdict(lambda: defaultdict(float)) 
    # Loop through paginated results
    epic_data = {}

    # Loop through paginated results
    while start_at < total_results:
        params = {
            "jql": jql_query,
            "fields": "summary, key, customfield_10000, subtasks, assignee",
            "startAt": start_at,
            "maxResults": max_results
        }
        
        response = requests.get(url, headers=headers, params=params)
        logging.debug(f"Request URL: {response.url}")

        if response.status_code == 200:
            data = response.json()
            total_results = data['total']
            issues = data.get('issues', [])

            if not issues and start_at == 0:
                logging.error("No issues found matching the JQL query.")
                return []            
            # Process each issue
            for issue in issues:  
                epic_summary = "No Summary"              
                epic = issue['fields'].get(epic_field, 'No Epic')
                if epic in epic_data:
                    epic_summary = epic_data[epic]
                else:            
                    epic_details = get_ticket_details(epic)                    
                    if epic_details:
                        epic_summary = epic_details['fields']['summary']
                    epic_summary= epic_summary.strip()
                    epic_data[epic] = epic_summary

                subtasks = issue['fields'].get('subtasks', [])
                if subtasks:
                    for sub_task in subtasks:
                        sub_task_key = sub_task['key']
                        subtask_details = get_ticket_details(sub_task_key)
                        if subtask_details:
                            subtask_assignee = subtask_details['fields']['assignee']['displayName'] if subtask_details['fields']['assignee'] else "Unassigned"
                            original_estimate = subtask_details['fields']['aggregatetimeestimate']                    
                            original_estimate_sp = original_estimate / 60 / 60 / 8 if original_estimate else 0                                                

                            try:
                                if team_data[epic_summary][subtask_assignee]:                                
                                    team_data[epic_summary][subtask_assignee] += original_estimate_sp
                            except KeyError as e:                            
                                team_data[epic_summary][subtask_assignee] = original_estimate_sp                
                else:
                    assignee = issue['fields']['assignee']['displayName'] if issue['fields']['assignee'] else "Unassigned"
                    original_estimate = issue['fields'].get('aggregatetimeestimate', 0)  # Time in seconds
                    original_estimate_sp = original_estimate / 60 / 60 / 8 if original_estimate else 0  # Convert to story points

                    # Update team data for the parent issue
                    try:
                        if team_data[epic_summary][assignee]:                        
                            team_data[epic_summary][assignee] += original_estimate_sp
                    except KeyError as ex:
                        team_data[epic_summary][assignee] = original_estimate_sp               

            # Update startAt to fetch the next set of results
            start_at += max_results
        else:
            logging.error(f"Failed to fetch tickets. Status code: {response.status_code}")
            logging.error(f"Response content: {response.text}")
            return []    
    return team_data


def update_planned_hours_in_excel(excel_file, sheet_name, jira_data):
    # Load the Excel workbook and select the sheet    
    wb = openpyxl.load_workbook(excel_file)
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name.strip()]
    else:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
        
    assignees = []
    header_row = sheet[1]  # First row contains the headers
    for col in range(2, len(header_row)):  # Start from column 2 (which is column C in Excel)
        assignee_name = header_row[col].value
        if assignee_name and assignee_name  not in ['Key', 'Summary','Actual', 'Planned', 'Team Member', 'SP']:
            assignees.append(assignee_name)
    
    print(f"Assignees found in the sheet: {assignees}")
    # Iterate over each row in the sheet
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        epic_link = row[0].value  # Epic link in the first column
        epic_name = row[1].value  # Epic name in the second column
                
        if not epic_link or not epic_name:
            print(f"Skipping row: Epic link or Epic name is missing. epic_link: {epic_link}, epic_name: {epic_name}")
            continue
        epic_data = jira_data.get(epic_name, {})        
        if epic_data:
            # Iterate through each assignee and update their "Planned" column
            for idx, assignee in enumerate(assignees, start=0):  # Start from the second column (first assignee)
                if assignee in epic_data:                    
                    planned_hours = epic_data[assignee]                    
                    # Calculate the correct column for the "Planned" based on the index
                    planned_column = 3 + (idx * 3) + 2                     
                    sheet.cell(row=row[0].row, column=planned_column, value=planned_hours)
        else:
            # If no data found for the epic, continue without making changes to the "Planned" columns
            print(f"No data found for epic {epic_link} (Epic name: {epic_name}), skipping update.")
    
    # Save the updated workbook
    wb.save(excel_file)


def main():    
    parser = argparse.ArgumentParser(description="Fetch Jira estimates and time spent for each team member.")
    parser.add_argument('--issue_type', help="Issue type (e.g., Sub-task)", required=True)
    parser.add_argument('--responsible_team', help="Responsible SAFe Team", required=True)
    parser.add_argument('--label', help="label", required=True)
    parser.add_argument('--file_name', help="label", required=True)
    parser.add_argument('--sheet_name', help="label", required=True)
    
    args = parser.parse_args()

    #Fetch estimates and time spent from Jira
    print(f"Fetching the Teams data is Started...")
    response_data = fetch_estimates_and_time_spent(args.issue_type, args.responsible_team, args.label)
    print(f"Fetching the Teams data is completed... {response_data}")
    if response_data:
        print("Updating of the Sheet Started...")
        update_planned_hours_in_excel(args.file_name, args.sheet_name, response_data)
        print("Updating of the Sheet Completed...")
    else:
        print("Invalid Tea Data from Jira, Unable to update the sheet")
if __name__ == "__main__":
    main()